# 3열이 비어 있으면 먼저 전체를 포스태깅하고, txt로 변환
# (표준어\t제주어\t표준어포스태깅\t제주어포스태깅\n)
# python generate_corpus [Copora Directory] [Output Directory]

import os
import sys
import openpyxl

from konlpy.tag import Komoran


def main(corpora, output):
    filelist = os.listdir(corpora)
    tagger_stan = Komoran()
    tagger_jeju = Komoran(userdic='userdic.txt') # TODO: If not userdic

    for file in filelist:
        book = openpyxl.load_workbook(os.path.join(corpora, file))
        sheet = book.get_sheet_by_name("Sheet")

        tagged = (bool(sheet.cell(row=1, column=3).value) and bool(sheet.cell(row=1, column=4).value))

        if not tagged:
            for sample in sheet.rows:
                index = sample[0].row
                try:
                    stan = sample[0].value
                    pos_stan = ' '.join(tagger_stan.morphs(stan))
                    jeju = sample[1].value
                    pos_jeju = ' '.join(tagger_jeju.morphs(jeju))
                except:
                    continue
                else:
                    sheet.cell(row=index, column=3).value = pos_stan
                    sheet.cell(row=index, column=4).value = pos_jeju

            book.save(os.path.join(corpora, file))

        filename = file[:file.find('.')]
        if not os.path.exists(output):
            os.makedirs(output)
        output_dir = os.path.join(output, filename + '.txt')    # Exception: Output Dir not Exists
        output_file = open(output_dir, 'w')

        for sample in sheet.rows:
            try:
                line = '\t'.join([s.value for s in sample[:5]]) + '\n'  # Exception: s.value can be no string
            except TypeError:
                continue
            else:
                output_file.write(line)

        output_file.close()
        book.close()


if __name__ == '__main__':
    args = sys.argv[1:]
    main(*args)
